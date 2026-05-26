#!/usr/bin/env python3
"""Convert an Excel file (.xlsx) to a self-contained HTML file.

Features:
  - Multiple sheets → tabbed HTML
  - Merged cells → colspan/rowspan + metadata
  - Formulas → displayed value + data-formula attribute
  - Column width auto-sizing
  - Clean, readable table styling
  - Embedded JSON metadata for writeback (html2excel.py)

Usage:
    python3 excel2html.py <input.xlsx> [output.html]

If output path is omitted, writes <input>.html next to the source file.
"""
import json
import os
import sys
from datetime import datetime
from copy import copy
from pathlib import Path

from runtime_support import configure_current_process_env


configure_current_process_env(Path(__file__).resolve().parents[1])

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("ERROR: openpyxl not installed. Run via run_python.py wrapper.", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell_ref(col_idx, row_idx):
    """Return 'A1'-style reference from 1-based col/row."""
    return f"{get_column_letter(col_idx)}{row_idx}"


def _cell_value_to_str(cell):
    """Convert a cell value to display string."""
    val = cell.value
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)


def _cell_type(cell, has_formula=False):
    """Determine the semantic type of a cell."""
    if has_formula:
        return "formula"
    val = cell.value
    if val is None:
        return "empty"
    if isinstance(val, bool):
        return "boolean"
    if isinstance(val, (int, float)):
        return "number"
    if isinstance(val, datetime):
        return "date"
    return "string"


def _escape_html(s):
    """Minimal HTML escape."""
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
         .replace('"', "&quot;")
    )


# ---------------------------------------------------------------------------
# Core conversion
# ---------------------------------------------------------------------------

def sheet_to_html(wb, sheet_name, sheet_idx):
    """Convert one worksheet to an HTML string (the <div class="sheet"> block)."""
    ws = wb[sheet_name]

    # --- Collect merge info ---
    # merge_cells gives us e.g. "A1:B3"
    merge_map = {}       # cell_ref → merge_range_str  (only on top-left cell)
    covered_cells = set()  # cells that are covered by a merge (not top-left)

    for merge_range in ws.merged_cells.ranges:
        rng_str = str(merge_range)  # e.g. "A1:B3"
        tl = merge_range.start_cell  # top-left cell
        merge_map[_cell_ref(tl.column, tl.row)] = rng_str

        # All cells in the range except top-left are "covered"
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                ref = _cell_ref(col, row)
                if ref != _cell_ref(tl.column, tl.row):
                    covered_cells.add(ref)

    # --- Collect formula info ---
    formulas = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == "f" and cell.value and str(cell.value).startswith("="):
                formulas[_cell_ref(cell.column, cell.row)] = str(cell.value)

    # --- Determine dimensions ---
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # --- Column widths (approximate: ~8px per char, min 60px) ---
    col_widths = {}
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        # openpyxl column_dimensions gives width in "characters"
        dim = ws.column_dimensions.get(letter)
        if dim and dim.width:
            col_widths[col_idx] = max(60, int(dim.width * 8))
        else:
            col_widths[col_idx] = 80

    # --- Build table ---
    rows_html = []

    # Header row (column letters)
    header_cells = []
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        w = col_widths.get(col_idx, 80)
        header_cells.append(f'<th style="min-width:{w}px">{letter}</th>')
    rows_html.append("        <tr>" + "".join(header_cells) + "</tr>")

    # Data rows
    for row_idx in range(1, max_row + 1):
        cells_html = []
        for col_idx in range(1, max_col + 1):
            ref = _cell_ref(col_idx, row_idx)

            if ref in covered_cells:
                # Hidden cell covered by a merge
                cells_html.append(
                    f'          <td data-cell="{ref}" data-covered="true" style="display:none"></td>'
                )
                continue

            cell = ws.cell(row=row_idx, column=col_idx)
            has_formula = ref in formulas
            formula_str = formulas.get(ref, "")
            cell_type = _cell_type(cell, has_formula)
            display_val = _cell_value_to_str(cell)
            raw_val = display_val

            # Build attributes
            attrs = [
                f'data-cell="{ref}"',
                f'data-type="{cell_type}"',
                f'data-raw="{_escape_html(raw_val)}"',
            ]
            if has_formula:
                attrs.append(f'data-formula="{_escape_html(formula_str)}"')

            # Merge info
            extra_style = ""
            if ref in merge_map:
                rng = merge_map[ref]
                merge_range_obj = None
                for mr in ws.merged_cells.ranges:
                    if str(mr) == rng:
                        merge_range_obj = mr
                        break
                if merge_range_obj:
                    colspan = merge_range_obj.max_col - merge_range_obj.min_col + 1
                    rowspan = merge_range_obj.max_row - merge_range_obj.min_row + 1
                    if colspan > 1:
                        attrs.append(f'colspan="{colspan}"')
                    if rowspan > 1:
                        attrs.append(f'rowspan="{rowspan}"')
                    attrs.append(f'data-merge="{rng}"')

            attr_str = " ".join(attrs)
            cells_html.append(f'          <td {attr_str}>{_escape_html(display_val)}</td>')

        rows_html.append("        <tr>" + "".join(cells_html) + "</tr>")

    # --- Compose the sheet div ---
    table_html = "\n".join(rows_html)
    merges_json = json.dumps(list(merge_map.values()))
    formulas_json = json.dumps(formulas)

    sheet_div = f"""
    <div class="sheet" data-sheet="{sheet_idx}" data-merges='{merges_json}' data-formulas='{formulas_json}'>
      <table>
        <thead>
          <tr><th class="row-num">#</th>{"".join(header_cells)}</tr>
        </thead>
        <tbody>
{table_html}
        </tbody>
      </table>
    </div>"""

    return sheet_div


# ---------------------------------------------------------------------------
# HTML assembly
# ---------------------------------------------------------------------------

HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{title}</title>
  <style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; background: #f5f5f5; color: #333; padding: 16px; }}
    h1 {{ font-size: 1.3rem; margin-bottom: 12px; color: #222; }}

    /* Tabs */
    .sheet-tabs {{ display: flex; gap: 4px; margin-bottom: 0; }}
    .sheet-tabs button {{
      padding: 8px 20px; border: 1px solid #ccc; border-bottom: none;
      background: #e8e8e8; cursor: pointer; font-size: 0.9rem; border-radius: 6px 6px 0 0;
      transition: background 0.15s;
    }}
    .sheet-tabs button:hover {{ background: #d0d0d0; }}
    .sheet-tabs button.active {{ background: #fff; font-weight: 600; border-bottom: 1px solid #fff; }}

    /* Table */
    .sheet {{ display: none; background: #fff; border: 1px solid #ccc; border-radius: 0 6px 6px 6px; overflow-x: auto; }}
    .sheet.active {{ display: block; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 0.85rem; }}
    th, td {{ border: 1px solid #d0d0d0; padding: 6px 10px; text-align: left; white-space: nowrap; }}
    th {{ background: #f0f0f0; font-weight: 600; position: sticky; top: 0; z-index: 1; }}
    th.row-num {{ width: 40px; text-align: center; color: #999; }}
    td[data-covered="true"] {{ display: none; }}
    td[data-type="formula"] {{ background: #f9f9e0; }}
    td[data-type="number"] {{ text-align: right; font-variant-numeric: tabular-nums; }}

    /* Hover */
    tr:hover td {{ background: #eef4ff; }}
    tr:hover td[data-type="formula"] {{ background: #eee8b0; }}

    /* Footer */
    .info {{ margin-top: 10px; font-size: 0.8rem; color: #888; }}
  </style>
</head>
<body>
  <h1>{title}</h1>
  {tabs_html}
  {sheets_html}
  <div class="info">{info}</div>

  <script>
    document.querySelectorAll('.sheet-tabs button').forEach(btn => {{
      btn.addEventListener('click', () => {{
        document.querySelectorAll('.sheet-tabs button').forEach(b => b.classList.remove('active'));
        document.querySelectorAll('.sheet').forEach(s => s.classList.remove('active'));
        btn.classList.add('active');
        document.querySelector('.sheet[data-sheet="' + btn.dataset.sheet + '"]').classList.add('active');
      }});
    }});
  </script>
</body>
</html>
"""


def excel_to_html(input_path, output_path=None):
    """Main conversion entry point."""
    if not os.path.isfile(input_path):
        print(f"ERROR: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(input_path, data_only=False)
    filename = os.path.basename(input_path)
    title = os.path.splitext(filename)[0]

    # Determine output path
    if output_path is None:
        output_path = os.path.splitext(input_path)[0] + ".html"

    # Build tabs and sheet divs
    sheet_names = wb.sheetnames
    tabs_html_parts = []
    sheets_html_parts = []

    for idx, name in enumerate(sheet_names):
        active = ' active' if idx == 0 else ''
        tabs_html_parts.append(
            f'  <div class="sheet-tabs">'
            f'<button class="tab{active}" data-sheet="{idx}">{_escape_html(name)}</button>'
            f'</div>' if idx == 0 else
            f'  <button class="tab{active}" data-sheet="{idx}">{_escape_html(name)}</button>'
        )
    # Fix: wrap tabs in a single div
    tabs_html = '  <div class="sheet-tabs">\n'
    for idx, name in enumerate(sheet_names):
        active = ' active' if idx == 0 else ''
        tabs_html += f'    <button class="tab{active}" data-sheet="{idx}">{_escape_html(name)}</button>\n'
    tabs_html += '  </div>'

    for idx, name in enumerate(sheet_names):
        sheet_div = sheet_to_html(wb, name, idx)
        sheets_html_parts.append(sheet_div)

    sheets_html = "\n".join(sheets_html_parts)

    # Info line
    total_sheets = len(sheet_names)
    info_parts = [f"Source: {_escape_html(filename)}", f"Sheets: {total_sheets}"]
    for idx, name in enumerate(sheet_names):
        ws = wb[name]
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        info_parts.append(f"  {name}: {rows} rows × {cols} cols")
    info = " | ".join(info_parts)

    html = HTML_TEMPLATE.format(
        title=_escape_html(title),
        tabs_html=tabs_html,
        sheets_html=sheets_html,
        info=info,
    )

    # Write output
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"[excel2html] Converted: {input_path} → {output_path}")
    print(f"[excel2html] Sheets: {total_sheets}, Names: {', '.join(sheet_names)}")

    wb.close()
    return output_path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: excel2html.py <input.xlsx> [output.html]", file=sys.stderr)
        sys.exit(1)

    in_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else None
    excel_to_html(in_path, out_path)
